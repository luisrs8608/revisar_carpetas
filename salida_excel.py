from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CAMPOS_SALIDA = [
    "Sede",
    "NombreOriginalCarpeta",
    "NombreNormalizadoCarpeta",
    "NombreOrdenadoCarpeta",
    "RutaCompleta",
    "EstaVacia",
    "CantidadItems",
    "CantidadArchivos",
    "CantidadSubcarpetas",
    "EstadoComparacion",
    "PorcentajeSimilitud",
    "FilaGoogleSheet",
    "FechaOriginalGoogleSheet",
    "NombreEncontradoGoogleSheet",
    "TieneMinutosInf",
    "ValorMinutosInf",
]


NOMBRES_HOJAS = [
    "Salida actual",
    "Con minutos",
    "Carpetas no encontradas",
    "Sin minutos ni carpeta",
]


ESTADOS_CARPETA_NO_ENCONTRADA = {
    "No encontrado",
    "Sin datos en Google Sheet",
}


COLOR_ENCABEZADO = "1F4E78"
COLOR_TEXTO_ENCABEZADO = "FFFFFF"
COLORES_ESTADO = {
    "Coincidencia exacta": "E2F0D9",
    "Coincidencia probable": "EAF4E2",
    "Posible coincidencia - revisar": "FFF2CC",
    "No encontrado": "FCE4D6",
    "Sin datos en Google Sheet": "FCE4D6",
    "Sin carpeta coincidente": "FCE4D6",
}


ANCHOS_COLUMNAS = {
    "Sede": 28,
    "NombreOriginalCarpeta": 30,
    "NombreNormalizadoCarpeta": 30,
    "NombreOrdenadoCarpeta": 30,
    "RutaCompleta": 55,
    "EstaVacia": 12,
    "CantidadItems": 16,
    "CantidadArchivos": 18,
    "CantidadSubcarpetas": 21,
    "EstadoComparacion": 32,
    "PorcentajeSimilitud": 21,
    "FilaGoogleSheet": 18,
    "FechaOriginalGoogleSheet": 25,
    "NombreEncontradoGoogleSheet": 34,
    "TieneMinutosInf": 18,
    "ValorMinutosInf": 18,
}


def asegurar_extension_xlsx(ruta: Path) -> Path:
    """Convierte configuraciones antiguas o sin extensión a `.xlsx`."""
    if ruta.suffix.lower() == ".xlsx":
        return ruta

    return ruta.with_suffix(".xlsx")


def separar_datos_hojas(
    datos: list[dict],
) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    """Devuelve la salida general y los tres subconjuntos solicitados."""
    con_minutos = [
        fila
        for fila in datos
        if fila.get("TieneMinutosInf") == "Sí"
    ]
    carpetas_no_encontradas = [
        fila
        for fila in datos
        if fila.get("RutaCompleta")
        and fila.get("EstadoComparacion")
        in ESTADOS_CARPETA_NO_ENCONTRADA
    ]
    sin_minutos_ni_carpeta = [
        fila
        for fila in datos
        if fila.get("EstadoComparacion") == "Sin carpeta coincidente"
        and fila.get("TieneMinutosInf") == "No"
        and fila.get("FilaGoogleSheet") not in (None, "")
        and not fila.get("RutaCompleta")
    ]

    return (
        datos,
        con_minutos,
        carpetas_no_encontradas,
        sin_minutos_ni_carpeta,
    )


def _agregar_hoja(
    workbook: Workbook,
    nombre: str,
    datos: list[dict],
    nombre_tabla: str,
) -> None:
    hoja = workbook.create_sheet(title=nombre)
    hoja.sheet_view.showGridLines = False
    hoja.freeze_panes = "A2"
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.print_title_rows = "1:1"
    hoja.sheet_view.zoomScale = 85

    hoja.append(CAMPOS_SALIDA)

    for fila in datos:
        hoja.append([
            fila.get(campo, "")
            for campo in CAMPOS_SALIDA
        ])

    encabezado = hoja[1]
    relleno_encabezado = PatternFill(
        fill_type="solid",
        fgColor=COLOR_ENCABEZADO,
    )

    for celda in encabezado:
        celda.fill = relleno_encabezado
        celda.font = Font(
            name="Arial",
            size=10,
            color=COLOR_TEXTO_ENCABEZADO,
            bold=True,
        )
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    hoja.row_dimensions[1].height = 34
    ultima_fila = max(hoja.max_row, 1)
    ultima_columna = get_column_letter(len(CAMPOS_SALIDA))
    hoja.auto_filter.ref = f"A1:{ultima_columna}{ultima_fila}"

    if datos:
        tabla = Table(
            displayName=nombre_tabla,
            ref=f"A1:{ultima_columna}{ultima_fila}",
        )
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja.add_table(tabla)

    indices = {
        campo: indice
        for indice, campo in enumerate(CAMPOS_SALIDA, start=1)
    }
    columnas_con_texto_extenso = {
        indices["NombreOriginalCarpeta"],
        indices["NombreNormalizadoCarpeta"],
        indices["NombreOrdenadoCarpeta"],
        indices["RutaCompleta"],
        indices["EstadoComparacion"],
        indices["NombreEncontradoGoogleSheet"],
    }
    columnas_enteras = {
        indices["CantidadItems"],
        indices["CantidadArchivos"],
        indices["CantidadSubcarpetas"],
        indices["FilaGoogleSheet"],
    }

    for columna, campo in enumerate(CAMPOS_SALIDA, start=1):
        letra = get_column_letter(columna)
        hoja.column_dimensions[letra].width = ANCHOS_COLUMNAS[campo]

    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
        for celda in fila:
            celda.font = Font(name="Arial", size=10)
            celda.alignment = Alignment(
                vertical="top",
                wrap_text=celda.column in columnas_con_texto_extenso,
            )

            if celda.column in columnas_enteras and celda.value != "":
                celda.number_format = "#,##0"

        celda_similitud = fila[indices["PorcentajeSimilitud"] - 1]
        if celda_similitud.value != "":
            celda_similitud.number_format = "0.00"

        celda_estado = fila[indices["EstadoComparacion"] - 1]
        color_estado = COLORES_ESTADO.get(celda_estado.value)
        if color_estado:
            celda_estado.fill = PatternFill(
                fill_type="solid",
                fgColor=color_estado,
            )


def exportar_excel(datos: list[dict], archivo_salida: Path) -> None:
    """Genera un libro Excel con la salida general y dos vistas filtradas."""
    archivo_salida = asegurar_extension_xlsx(archivo_salida)
    archivo_salida.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    conjuntos = separar_datos_hojas(datos)
    nombres_tablas = [
        "TablaSalidaActual",
        "TablaConMinutos",
        "TablaCarpetasNoEncontradas",
        "TablaSinMinutosNiCarpeta",
    ]

    for nombre, filas, nombre_tabla in zip(
        NOMBRES_HOJAS,
        conjuntos,
        nombres_tablas,
    ):
        _agregar_hoja(workbook, nombre, filas, nombre_tabla)

    workbook.save(archivo_salida)
