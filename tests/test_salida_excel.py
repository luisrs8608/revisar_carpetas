import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from salida_excel import (
    CAMPOS_SALIDA,
    NOMBRES_HOJAS,
    asegurar_extension_xlsx,
    exportar_excel,
    separar_datos_hojas,
)


def crear_fila(**cambios):
    fila = {
        campo: ""
        for campo in CAMPOS_SALIDA
    }
    fila.update({
        "Sede": "SEDE CENTRO",
        "NombreOriginalCarpeta": "Ana_Perez",
        "NombreNormalizadoCarpeta": "ana perez",
        "NombreOrdenadoCarpeta": "ana perez",
        "RutaCompleta": "/sedes/SEDE CENTRO/Ana_Perez",
        "EstaVacia": "No",
        "CantidadItems": 3,
        "CantidadArchivos": 3,
        "CantidadSubcarpetas": 0,
        "EstadoComparacion": "Coincidencia exacta",
        "PorcentajeSimilitud": 100,
        "FilaGoogleSheet": 12,
        "FechaOriginalGoogleSheet": "1/6",
        "NombreEncontradoGoogleSheet": "Ana Perez",
        "TieneMinutosInf": "No",
        "ValorMinutosInf": "",
    })
    fila.update(cambios)
    return fila


class SepararDatosHojasTests(unittest.TestCase):
    def test_separa_minutos_y_carpetas_no_encontradas(self):
        salida = [
            crear_fila(TieneMinutosInf="Sí", ValorMinutosInf="30"),
            crear_fila(
                NombreOriginalCarpeta="Persona_Sin_Fila",
                EstadoComparacion="No encontrado",
            ),
            crear_fila(
                NombreOriginalCarpeta="Persona_Revisar",
                EstadoComparacion="Posible coincidencia - revisar",
            ),
            crear_fila(
                RutaCompleta="",
                EstadoComparacion="Sin carpeta coincidente",
                TieneMinutosInf="No",
            ),
        ]

        (
            general,
            con_minutos,
            no_encontradas,
            sin_minutos_ni_carpeta,
        ) = separar_datos_hojas(salida)

        self.assertIs(general, salida)
        self.assertEqual(len(con_minutos), 1)
        self.assertEqual(len(no_encontradas), 1)
        self.assertEqual(
            no_encontradas[0]["NombreOriginalCarpeta"],
            "Persona_Sin_Fila",
        )
        self.assertEqual(len(sin_minutos_ni_carpeta), 1)
        self.assertEqual(
            sin_minutos_ni_carpeta[0]["EstadoComparacion"],
            "Sin carpeta coincidente",
        )


class ExportarExcelTests(unittest.TestCase):
    def test_normaliza_la_extension_de_salida(self):
        self.assertEqual(
            asegurar_extension_xlsx(Path("resultado.csv")),
            Path("resultado.xlsx"),
        )
        self.assertEqual(
            asegurar_extension_xlsx(Path("resultado")),
            Path("resultado.xlsx"),
        )
        self.assertEqual(
            asegurar_extension_xlsx(Path("resultado.xlsx")),
            Path("resultado.xlsx"),
        )

    def test_genera_cuatro_hojas_con_datos_y_formato(self):
        datos = [
            crear_fila(TieneMinutosInf="Sí", ValorMinutosInf="30"),
            crear_fila(
                NombreOriginalCarpeta="Persona_Sin_Fila",
                EstadoComparacion="No encontrado",
                PorcentajeSimilitud=42.5,
            ),
            crear_fila(
                NombreOriginalCarpeta="Persona_Revisar",
                EstadoComparacion="Posible coincidencia - revisar",
                PorcentajeSimilitud=80,
            ),
            crear_fila(
                Sede="",
                NombreOriginalCarpeta="",
                RutaCompleta="",
                EstadoComparacion="Sin carpeta coincidente",
                PorcentajeSimilitud="",
                FilaGoogleSheet=18,
                NombreEncontradoGoogleSheet="Persona Pendiente",
                TieneMinutosInf="No",
                ValorMinutosInf="",
            ),
        ]

        with tempfile.TemporaryDirectory() as temporal:
            archivo = Path(temporal) / "resultado.xlsx"
            exportar_excel(datos, archivo)
            workbook = load_workbook(archivo)

            self.assertEqual(workbook.sheetnames, NOMBRES_HOJAS)
            self.assertEqual(workbook["Salida actual"].max_row, 5)
            self.assertEqual(workbook["Con minutos"].max_row, 2)
            self.assertEqual(
                workbook["Carpetas no encontradas"].max_row,
                2,
            )
            self.assertEqual(
                workbook["Sin minutos ni carpeta"].max_row,
                2,
            )

            for hoja in workbook.worksheets:
                self.assertEqual(
                    [celda.value for celda in hoja[1]],
                    CAMPOS_SALIDA,
                )
                self.assertEqual(hoja.freeze_panes, "A2")
                self.assertIsNotNone(hoja.auto_filter.ref)
                self.assertFalse(hoja.sheet_view.showGridLines)
                self.assertEqual(
                    hoja["A1"].fill.fgColor.rgb,
                    "001F4E78",
                )

            salida = workbook["Salida actual"]
            columna_items = CAMPOS_SALIDA.index("CantidadItems") + 1
            columna_similitud = (
                CAMPOS_SALIDA.index("PorcentajeSimilitud") + 1
            )
            self.assertEqual(salida.cell(2, columna_items).value, 3)
            self.assertEqual(
                salida.cell(3, columna_similitud).value,
                42.5,
            )

    def test_conserva_las_cuatro_hojas_si_no_hay_resultados(self):
        with tempfile.TemporaryDirectory() as temporal:
            archivo = Path(temporal) / "sin_datos.xlsx"
            exportar_excel([], archivo)
            workbook = load_workbook(archivo)

            self.assertEqual(workbook.sheetnames, NOMBRES_HOJAS)
            self.assertTrue(all(
                hoja.max_row == 1
                for hoja in workbook.worksheets
            ))


if __name__ == "__main__":
    unittest.main()
